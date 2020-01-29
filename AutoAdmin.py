from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from io import StringIO 
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os
import sys
import pathlib
import ctypes
import datetime
import config
import openpyxl
import re 
import smtplib, ssl 


class AutoAdmin:
    
    def __init__(self):
        print("AutoAdmin started")

    def convert_pdf_to_txt(self, path):
        rsrcmgr = PDFResourceManager()
        retstr = StringIO()
        codec = 'utf-8'
        laparams = LAParams()
        device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
        fp = open(path, 'rb')
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        password = ""
        maxpages = 1
        caching = True
        pagenos=set()
        for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, 
                                      check_extractable=True):
            interpreter.process_page(page)
        fp.close()
        device.close()

        str = retstr.getvalue() # raw
        retstr.close()
    
        # Categorize docs
        if "DOCUMENT TYPE 1" in str:
            client_status = "STATUS1"
        elif "DOCUMENT TYPE 2" in str:
            client_status = "STATUS2"
        elif "DOCUMENT TYPE 3" in str:
            client_status = "STATUS3"
        else:
            execute_alert('Warning', "The type of document is not identified", 0)
            exit()

        client_app_num = re.search(config.EXCEL_APP_NUM_REGEX, str).group()
        original_date = self.search_date(str)
        date_in_format = self.convert_date(original_date)
    
        return {"num": client_app_num, "status": client_status, "date": date_in_format, "path": path}

    

    def execute_alert(self, title, text, style):
        return ctypes.windll.user32.MessageBoxW(0, text, title, style)


    def modify_excel(self, client_info):
        sheets = openpyxl.load_workbook(config.EXCEL_FILE_PATH)
        selected_sheet = sheets[client_info["sheet"]]

        if "STATUS1" in client_info["status"]:
            client_info["note"] = client_info["note"] + "\n" + \
                                  client_info["date"] + ": " + "STATUS1 letter"
        elif "STATUS2" in client_info["status"]:
            client_info["note"] = client_info["note"] + "\n" + \
                                  client_info["date"] + ": " + "STATUS2 letter"
        elif "STATUS3" in client_info["status"]:
            client_info["note"] = client_info["note"] + "\n" + \
                                  client_info["date"] + ": " + "STATUS3 letter"
        else:
            execute_alert('Not Found', "The status of client's application is not detected." + \
                                       "Please process it manually", 0)
            exit()

        selected_sheet.cell(row=client_info["row"], column=config.EXCEL_NOTE).value = client_info["note"]
        sheets.save(config.EXCEL_FILE_PATH.resolve())


    def modify_file_name(self, client_info): 

        last_name = client_info["last"].split(" ")[0]
        first_name = client_info["first"].split(" ")[0]
    
        if "STATUS1" in client_info["status"]:
            new_file_name = last_name + ", " + first_name + " - STATUS1 letter"
        elif "STATUS2" in client_info["status"]:
            new_file_name = last_name + ", " + first_name + " - STATUS2 letter"
        elif "STATUS3" in client_info["status"]:
            new_file_name = last_name + ", " + first_name + " - STATUS3 letter"
        else:
            execute_alert('Not Found', "The status of client's application is not detected. " + \
                                       "Please process it manually", 0)
            exit()
        new_file_name = new_file_name.lower() + ".pdf"
    
        client_info["path"].rename(pathlib.Path(pathlib.Path.cwd(), new_file_name))
        client_info["file_name"] = new_file_name

        return client_info
    

    def search_date(self, str):
        years = r'((?:19|20)\d\d)'
        pattern = r'(%%s) +(%%s), *%s' % years

        thirties = pattern % (
            "September|April|June|November",
            r'0?[1-9]|[12]\d|30')

        thirtyones = pattern % (
            "January|March|May|July|August|October|December",
            r'0?[1-9]|[12]\d|3[01]')

        fours = '(?:%s)' % '|'.join('%02d' % x for x in range(4, 100, 4))

        feb = r'(February) +(?:%s|%s)' % (
            r'(?:(0?[1-9]|1\d|2[0-8])), *%s' % years, # 1-28 any year
            r'(?:(29), *((?:(?:19|20)%s)|2000))' % fours)  # 29 leap years only

        config.EXCEL_DATE_REGEX = '|'.join('(?:%s)' % x for x in (thirties, thirtyones, feb))

        return re.search(config.EXCEL_DATE_REGEX, str).group()
    

    def convert_date(self, dateInStr):
    
        dateInList = dateInStr.split(" ")
    
        if "January" in dateInStr:
            month = "01"
        elif "February" in dateInStr:
            month = "02"
        elif "March" in dateInStr:
            month = "03"
        elif "April" in dateInStr:
            month = "04"
        elif "May" in dateInStr:
            month = "05"
        elif "June" in dateInStr:
            month = "06"
        elif "July" in dateInStr:
            month = "07"
        elif "August" in dateInStr:
            month = "08"
        elif "September" in dateInStr:
            month = "09"
        elif "October" in dateInStr:
            month = "10"
        elif "November" in dateInStr:
            month = "11"
        elif "December" in dateInStr:
            month = "12"
        else:
            execute_alert('Date Not Found', "Cannot find a date in the document", 0)
            exit()
        return dateInList[2] + "-" + month + "-" + dateInList[1][:-1]



    def read_excel(self, client_num):
        sheets = openpyxl.load_workbook(config.EXCEL_FILE_PATH)
        sheetNames = sheets.sheetnames
        for sheet_name in config.EXCEL_SHEETS:
            selectedSheet = sheets[sheet_name]
            for r in selectedSheet.iter_rows(min_row=config.EXCEL_FIRST_ROW, max_row=config.EXCEL_LAST_ROW, 
                                             min_col=config.EXCEL_APP_NUM, max_col=config.EXCEL_APP_NUM):
                for cell in r:
                    client_row = cell.row
                    if cell.value == client_num:
                        config.EXCEL_SELECTED_SHEET = selectedSheet

                        client_last = config.EXCEL_SELECTED_SHEET.cell(row=client_row, 
                                                                       column=config.EXCEL_LAST_NAME).value
                        client_first = config.EXCEL_SELECTED_SHEET.cell(row=client_row, 
                                                                        column=config.EXCEL_FIRST_NAME).value
                        client_nationality = config.EXCEL_SELECTED_SHEET.cell(row=client_row, 
                                                                              column=config.EXCEL_NATIONALITY).value
                        client_type = config.EXCEL_SELECTED_SHEET.cell(row=client_row, 
                                                                       column=config.EXCEL_TYPE).value
                        client_agent = config.EXCEL_SELECTED_SHEET.cell(row=client_row, 
                                                                        column=config.EXCEL_AGENT_NAME).value

                        client_note = config.EXCEL_SELECTED_SHEET.cell(row=client_row, 
                                                                       column=config.EXCEL_NOTE).value
                        if not client_note:
                            client_note = ''
                            print("FINISHED read_excel")
                        return {"last": client_last, "first": client_first, "nationality": client_nationality, 
                                "type": client_type, "agent": client_agent, "row": client_row, 
                                "note": client_note, "sheet": sheet_name}

                    elif cell.row >= config.EXCEL_LAST_ROW:
                        client_row = None
                        break
                else:
                    continue
                break

            execute_alert('Not Found', "Client's application number is not found. Please process it manually", 0)
            exit()

 

    def tuple_to_str(self, message):
        rst = ''.join(message)

        return rst

    def create_email(self, last, first, nationality, agent, type, status):
        if "TYPE1" in type:
            type_for_msg = "TYPE 1"
        elif "TYPE2" in type:
            type_for_msg = "TYPE 2"
        elif "TYPE3" in type:
            type_for_msg = "TYPE 3"
        else:
            execute_alert('Not Found', "The type of client's application is not detected.", 0)
            exit()

        if status == "STATUS1":
            result = ("Hello ",
                      agent,
                      ",\n\n",
                      "Please find the attached TYPE 1 letter from ORGANIZATION NAME regarding ",
                      first, " ", last,
                      "'s ",
                      type_for_msg, " application. \n\n",
                      "We will notify you accordingly. \n\nBest regards,\n",
                      config.GOOGLE_AGENT_NAME)

        elif status == "STATUS2":
            result = ("Hello ",
                      agent,
                      ",\n\n",
                      "Please find the attached STATUS2 letter from ORGANIZATION NAME regarding ",
                      first, " ", last,
                      "'s ",
                      type_for_msg, " application. \n\n",
                      "We will notify you accordingly. \n\nBest regards,\n",
                      config.GOOGLE_AGENT_NAME)

        elif status == "STATUS3":
            result = ("Hello ",
                      agent,
                      ",\n\n",
                      "Please find the attached STATUS3 letter from ORGANIZATION NAME regarding ",
                      first, " ", last,
                      "'s ",
                      type_for_msg, " application. \n\n",
                      "We will notify you accordingly. \n\nBest regards,\n",
                      config.GOOGLE_AGENT_NAME)

        else:
            execute_alert('Not Found', "The status of client's application is not detected. Please process it manually", 0)
            exit()
   
        print("FINISHED create_email")
        result = tuple_to_str(result)
        return result    


    def send_email(self, client_info):
        print("STARTED send_email")

        text = create_email(client_info["last"], client_info["first"], client_info["nationality"], 
                            client_info["agent"], client_info["type"], client_info["status"])

        port = 587  # For SSL
        smtp_server = "smtp.gmail.com" # Smtp
        sender_email = GOOGLE_EMAIL  # Enter your address
        receiver_email = "PLACEHOLDER"  # Enter receiver address
 
    
        if (client_info["agent"] in DEPT1_AGENT):
            receiver_email = GOOGLE_DEPT1
        elif (client_info["agent"] in DEPT2_AGENT):
            receiver_email = GOOGLE_DEPT2
        elif (client_info["agent"] in DEPT3_AGENT):
            receiver_email = GOOGLE_DEPT3
        else:
            execute_alert('Not Found', "Unidentified agent. Email was not sent and Excel was not updated. " + \
                                       "Please check whether the agent's name and email address are added in a list", 0)
            exit()


        message = MIMEMultipart()
        message["From"] = GOOGLE_EMAIL
        message["To"] = "RECEIVER_EMAIL"
        message["Subject"] = "[" + client_info['status'].split(' ')[0] + "] " + client_info['last'] + \
                             ", " + client_info['first'] + " (" + client_info['type'] + ") - " + \
                             client_info['nationality']

        message.attach(MIMEText(text, "plain"))

        with open(client_info["path"].resolve(), "rb") as attachment:

            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
            "Content-Disposition",
            f"attachment; filename= {client_info['file_name']}",
        )
        message.attach(part)

        text = message.as_string()
        context = ssl.create_default_context()

        with smtplib.SMTP(smtp_server, port) as server:
            try:
                server = smtplib.SMTP(smtp_server, port)
                server.ehlo()
                server.starttls(context=context)
                server.ehlo()
                server.login(GOOGLE_EMAIL, GOOGLE_PW)
                server.sendmail(GOOGLE_EMAIL, "RECEIVER_EMAIL", text) 
            except Exception as e:
                print(e)
                exit()
            finally:
                server.quit()

        return 0


if __name__ == "__main__":
    path ='%s' %sys.argv[1]  
    path = pathlib.Path(pathlib.Path.cwd(), path)
    
    obj = AutoAdmin()

    client_info = obj.convert_pdf_to_txt(path)
    
    more_client_info = obj.read_excel(client_info["num"])
   
    client_info.update(more_client_info)

    client_info = obj.modify_file_name(client_info)

    obj.end_email(client_info)
    obj.modify_excel(client_info)

    input('Press ENTER to exit')
    
